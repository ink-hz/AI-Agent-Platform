from __future__ import annotations

import json
import re
from collections.abc import Iterable
from io import BytesIO
from uuid import UUID
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
)

from .panorama_models import PanoramaReport, PublicJobSnapshot, TalentSource, thaw_json

_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_FONT_NAME = "STSong-Light"
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")
_CAMPUS_TEXT = re.compile(
    r"校招|校园招聘|应届|毕业生|实习|(?<![a-z])(?:campus|graduate|intern(?:ship)?)(?![a-z])",
    re.IGNORECASE,
)
_CAMPUS_URL = re.compile(
    r"(?:^|[/_.-])(?:campus|graduate|intern(?:ship|recruitment)?)(?:$|[/_.?&#-])",
    re.IGNORECASE,
)
_SOCIAL = re.compile(
    r"社招|社会招聘|社会人才|(?<![a-z])experienced(?![a-z])|professional-hire",
    re.IGNORECASE,
)
_DIRECTION_PATTERNS = (
    ("algorithm", "算法", re.compile(r"算法|人工智能|机器学习|视觉|点云|(?<![a-z])ai(?![a-z])|slam", re.IGNORECASE)),
    ("optics", "光学", re.compile(r"光学|镜头|zemax|成像", re.IGNORECASE)),
    ("hardware", "硬件", re.compile(r"硬件|电子|电路|pcb|嵌入式", re.IGNORECASE)),
    ("structure", "结构", re.compile(r"结构|机械|机电|模具|cad", re.IGNORECASE)),
    ("software", "软件", re.compile(r"软件|前端|后端|客户端|java|c\+\+|python", re.IGNORECASE)),
    ("manufacturing", "制造工艺", re.compile(r"制造|工艺|质量|dqe|生产|供应链", re.IGNORECASE)),
)


def _text(value: object, maximum: int = 32767) -> str:
    selected = _CONTROL.sub("", str(value))
    return selected[:maximum]


def _xlsx_text(value: object, maximum: int = 32767) -> str:
    selected = _text(value, maximum)
    return f"'{selected[: maximum - 1]}" if selected.startswith(_FORMULA_PREFIXES) else selected


def recruitment_track(item: PublicJobSnapshot) -> str:
    text = f"{item.title} {item.duty_excerpt} {item.requirement_excerpt}"
    if _CAMPUS_TEXT.search(text) or _CAMPUS_URL.search(item.source_url):
        return "campus"
    if _SOCIAL.search(f"{text} {item.source_url}"):
        return "social"
    return "unknown"


def technical_direction(item: PublicJobSnapshot) -> str:
    text = f"{item.title} {item.duty_excerpt} {item.requirement_excerpt}"
    for key, _, pattern in _DIRECTION_PATTERNS:
        if pattern.search(text):
            return key
    return "other"


def _direction_label(item: PublicJobSnapshot) -> str:
    key = technical_direction(item)
    return next((label for value, label, _ in _DIRECTION_PATTERNS if value == key), "其他")


def _track_label(item: PublicJobSnapshot) -> str:
    return {"social": "社招", "campus": "校招/实习", "unknown": "待分类"}[
        recruitment_track(item)
    ]


def filter_panorama_snapshots(
    report: PanoramaReport,
    *,
    source_id: UUID | None = None,
    recruitment_track_filter: str | None = None,
    location: str | None = None,
    status: str | None = None,
    technical_direction_filter: str | None = None,
) -> tuple[PublicJobSnapshot, ...]:
    return tuple(
        item
        for item in report.snapshots
        if (source_id is None or item.source_id == source_id)
        and (
            recruitment_track_filter is None
            or recruitment_track(item) == recruitment_track_filter
        )
        and (location is None or item.location == location)
        and (status is None or item.status == status)
        and (
            technical_direction_filter is None
            or technical_direction(item) == technical_direction_filter
        )
    )


def _channel_snapshots(
    source: TalentSource, snapshots: Iterable[PublicJobSnapshot]
) -> dict[str, list[PublicJobSnapshot]]:
    result = {url: [] for url in source.approved_urls}
    for snapshot in snapshots:
        matches = [
            url
            for url in source.approved_urls
            if snapshot.source_url == url
            or snapshot.source_url.startswith(url.rstrip("/") + "/")
        ]
        if matches:
            result[max(matches, key=len)].append(snapshot)
    return result


def _source_names(report: PanoramaReport) -> dict[object, str]:
    return {source.source_id: source.canonical_name for source in report.sources}


def _scoped_content(
    report: PanoramaReport,
    snapshots: tuple[PublicJobSnapshot, ...],
    *,
    filtered: bool,
):
    if not filtered:
        return (
            report.sources,
            report.insight.facts,
            report.insight.inferences,
            report.insight.unknowns,
            report.insight.summary,
            thaw_json(report.insight.direction_clusters),
        )
    source_ids = {item.source_id for item in snapshots}
    snapshot_ids = {str(item.snapshot_id) for item in snapshots}
    sources = tuple(source for source in report.sources if source.source_id in source_ids)
    facts = tuple(
        fact
        for fact in report.insight.facts
        if str(fact["snapshot_id"]) in snapshot_ids
    )
    fact_ids = {str(fact["fact_id"]) for fact in facts}
    inferences = tuple(
        item
        for item in report.insight.inferences
        if all(str(value) in fact_ids for value in item["basis_fact_ids"])
    )
    directions: dict[str, int] = {}
    for item in snapshots:
        label = _direction_label(item)
        directions[label] = directions.get(label, 0) + 1
    return (
        sources,
        facts,
        inferences,
        (),
        f"筛选结果：{len(snapshots)} 条岗位，覆盖 {len(sources)} 家公司。",
        directions,
    )


def _fit_columns(sheet, *, maximum: int = 48) -> None:
    for index, column in enumerate(sheet.columns, start=1):
        width = min(
            maximum, max((len(str(cell.value or "")) for cell in column), default=8) + 2
        )
        sheet.column_dimensions[get_column_letter(index)].width = max(10, width)


def _sheet(
    workbook: Workbook,
    title: str,
    headers: tuple[str, ...],
    rows: list[tuple[object, ...]],
):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(tuple(_xlsx_text(value) for value in row))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="185FA9")
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _fit_columns(sheet)
    return sheet


def build_panorama_xlsx(
    report: PanoramaReport,
    snapshots: tuple[PublicJobSnapshot, ...] | None = None,
    *,
    filtered: bool = False,
) -> bytes:
    names = _source_names(report)
    selected_snapshots = report.snapshots if snapshots is None else snapshots
    sources, facts, inferences, unknowns, summary, directions = _scoped_content(
        report, selected_snapshots, filtered=filtered
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    overview = workbook.create_sheet("总览")
    overview.append(("全景分析版本", report.insight.version_number))
    overview.append(("分析时间", _xlsx_text(report.insight.created_at.isoformat())))
    overview.append(
        ("覆盖公司", _xlsx_text("、".join(source.canonical_name for source in sources)))
    )
    overview.append(("公开岗位记录", len(selected_snapshots)))
    overview.append(("摘要", _xlsx_text(summary)))
    overview.append(
        (
            "研发方向",
            _xlsx_text(
                json.dumps(
                    directions,
                    ensure_ascii=False,
                    sort_keys=True,
                )
            ),
        )
    )
    overview.column_dimensions["A"].width = 18
    overview.column_dimensions["B"].width = 90
    for cell in overview["A"]:
        cell.font = Font(bold=True, color="17345C")
    for row in overview.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _sheet(
        workbook,
        "岗位明细",
        ("公司", "岗位", "招聘类型", "技术方向", "地点", "状态", "职责", "要求", "来源", "观测时间"),
        [
            (
                names.get(item.source_id, "关注公司"),
                item.title,
                _track_label(item),
                _direction_label(item),
                item.location,
                item.status,
                item.duty_excerpt,
                item.requirement_excerpt,
                item.source_url,
                item.observed_at.isoformat(),
            )
            for item in selected_snapshots
        ],
    )
    _sheet(
        workbook,
        "情报来源",
        ("公司", "渠道地址", "岗位命中数", "最近观测", "证据状态"),
        [
            (
                source.canonical_name,
                url,
                len(items),
                max((item.observed_at for item in items), default=None).isoformat()
                if items
                else "",
                "已形成岗位证据" if items else "未形成岗位证据，待确认",
            )
            for source in sources
            for url, items in _channel_snapshots(source, selected_snapshots).items()
        ],
    )
    _sheet(
        workbook,
        "公开事实",
        ("事实", "来源", "观测时间"),
        [
            (fact["text"], fact["source_url"], fact["observed_at"])
            for fact in facts
        ],
    )
    _sheet(
        workbook,
        "AI推断",
        ("推断", "依据事实ID"),
        [
            (item["text"], "、".join(str(value) for value in item["basis_fact_ids"]))
            for item in inferences
        ],
    )
    _sheet(
        workbook,
        "待确认",
        ("未知项",),
        [(item["text"],) for item in unknowns],
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _paragraph(value: object, style: ParagraphStyle) -> Paragraph:
    selected = _CONTROL.sub("", str(value))
    return Paragraph(escape(selected).replace("\n", "<br/>"), style)


def _footer(canvas, document) -> None:
    canvas.saveState()
    canvas.setFont(_FONT_NAME, 8)
    canvas.setFillColor(colors.HexColor("#6B7F93"))
    canvas.drawRightString(
        A4[0] - 18 * mm, 9 * mm, f"Orbbec HR Agent | 第 {document.page} 页"
    )
    canvas.restoreState()


def build_panorama_pdf(
    report: PanoramaReport,
    snapshots: tuple[PublicJobSnapshot, ...] | None = None,
    *,
    filtered: bool = False,
) -> bytes:
    pdfmetrics.registerFont(UnicodeCIDFont(_FONT_NAME))
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=17 * mm,
        bottomMargin=17 * mm,
        title=f"招聘情报全景分析 - 第 {report.insight.version_number} 版",
        author="Orbbec HR Agent",
    )
    base = getSampleStyleSheet()
    title = ParagraphStyle(
        "ChineseTitle",
        parent=base["Title"],
        fontName=_FONT_NAME,
        fontSize=22,
        leading=30,
        textColor=colors.HexColor("#17345C"),
        alignment=TA_CENTER,
    )
    heading = ParagraphStyle(
        "ChineseHeading",
        parent=base["Heading2"],
        fontName=_FONT_NAME,
        fontSize=15,
        leading=21,
        textColor=colors.HexColor("#185FA9"),
        spaceBefore=10,
        spaceAfter=7,
    )
    body = ParagraphStyle(
        "ChineseBody",
        parent=base["BodyText"],
        fontName=_FONT_NAME,
        fontSize=9.5,
        leading=15,
        textColor=colors.HexColor("#203A54"),
        wordWrap="CJK",
    )
    small = ParagraphStyle(
        "ChineseSmall",
        parent=body,
        fontSize=8,
        leading=12,
        textColor=colors.HexColor("#536B84"),
    )
    names = _source_names(report)
    selected_snapshots = report.snapshots if snapshots is None else snapshots
    sources, facts, inferences, unknowns, summary, directions = _scoped_content(
        report, selected_snapshots, filtered=filtered
    )
    story = [
        _paragraph(f"招聘情报全景分析 - 第 {report.insight.version_number} 版", title),
        Spacer(1, 7 * mm),
        _paragraph(summary, body),
        Spacer(1, 4 * mm),
        _paragraph(
            f"覆盖公司：{'、'.join(source.canonical_name for source in sources) or '无匹配公司'}",
            small,
        ),
        _paragraph(
            f"分析时间：{report.insight.created_at.isoformat()} | 公开岗位记录：{len(selected_snapshots)}",
            small,
        ),
        _paragraph("研发与业务方向", heading),
        _paragraph(
            json.dumps(
                directions,
                ensure_ascii=False,
                sort_keys=True,
            ),
            body,
        ),
        _paragraph("公开事实", heading),
    ]
    for fact in facts:
        story.extend(
            (
                _paragraph(f"- {fact['text']}", body),
                _paragraph(
                    f"来源：{fact['source_url']} | 观测：{fact['observed_at']}", small
                ),
                Spacer(1, 2 * mm),
            )
        )
    story.append(_paragraph("AI 推断", heading))
    story.extend(
        _paragraph(
            f"- {item['text']}（依据：{'、'.join(str(value) for value in item['basis_fact_ids'])}）",
            body,
        )
        for item in inferences
    )
    story.append(_paragraph("仍待确认", heading))
    story.extend(
        _paragraph(f"- {item['text']}", body) for item in unknowns
    )
    story.extend((PageBreak(), _paragraph("岗位明细", heading)))
    if not selected_snapshots:
        story.append(_paragraph("当前筛选条件下没有岗位记录。", body))
    for item in selected_snapshots:
        story.extend(
            (
                _paragraph(
                    f"{names.get(item.source_id, '关注公司')} | {item.title}", body
                ),
                _paragraph(
                    f"{_track_label(item)} | {_direction_label(item)} | {item.location} | {item.status}",
                    small,
                ),
                _paragraph(f"职责：{item.duty_excerpt}", small),
                _paragraph(f"要求：{item.requirement_excerpt}", small),
                _paragraph(
                    f"来源：{item.source_url} | 观测：{item.observed_at.isoformat()}",
                    small,
                ),
                Spacer(1, 3 * mm),
            )
        )
    story.extend((Spacer(1, 5 * mm), _paragraph("情报来源矩阵", heading)))
    for source in sources:
        story.append(_paragraph(source.canonical_name, body))
        story.extend(_paragraph(f"- {url}", small) for url in source.approved_urls)
        story.append(Spacer(1, 3 * mm))
    document.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return output.getvalue()


__all__ = [
    "build_panorama_pdf",
    "build_panorama_xlsx",
    "filter_panorama_snapshots",
    "recruitment_track",
    "technical_direction",
]
